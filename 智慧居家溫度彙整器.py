import os
import pandas as pd
import numpy as np
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from typing import Optional, List, Tuple, Dict, Any

# ============================
# 單檔處理函式
# ----------------------------

def process_file(file_path: str, room_name: str) -> Tuple:
    """處理單個 Excel 文件的溫度數據，篩選指定房間，計算平均溫度和最高溫度"""
    try:
        data = pd.read_excel(file_path, sheet_name='Room Data')
    except ValueError as e:
        print(f"Error reading file {file_path}: {e}")
        return tuple([None] * 4)  # 返回 4 個值：平均溫度、最高溫度、時間範圍、記錄數
    
    # 檢查必要欄位
    if 'Room' not in data.columns or 'Temperature' not in data.columns or 'Time' not in data.columns:
        print("No Room, Temperature, or Time column")
        return tuple([None] * 4)
    
    # 篩選指定房間的數據
    filtered_data = data[data['Room'] == room_name]
    if filtered_data.empty:
        return tuple([None] * 4)
    
    # 計算平均溫度和最高溫度
    avg_temp = filtered_data['Temperature'].mean()
    max_temp = filtered_data['Temperature'].max()
    time_range = f"{filtered_data['Time'].min()} - {filtered_data['Time'].max()}"
    record_count = len(filtered_data)
    
    return (avg_temp, max_temp, time_range, record_count)

# ============================
# 輔助函式
# ----------------------------

def extract_room_name(file_name: str) -> Optional[str]:
    """從文件名提取房間名稱"""
    match = re.search(r'(LivingRoom|Bedroom|Kitchen)_', file_name, re.IGNORECASE)
    return match.group(1) if match else None

def extract_day_number(file_name: str) -> Optional[int]:
    """從文件名提取日期編號"""
    match = re.search(r'Day(\d+)', file_name, re.IGNORECASE)
    return int(match.group(1)) if match else None

def color_based_on_room(room: str) -> Optional[PatternFill]:
    """根據房間返回對應的顏色填充"""
    color_map = {
        'LivingRoom': 'FFCCCB',  # 淺紅
        'Bedroom': 'ADD8E6',    # 淺藍
        'Kitchen': '90EE90'     # 淺綠
    }
    return PatternFill(start_color=color_map.get(room), end_color=color_map.get(room), fill_type='solid') if room in color_map else None

def check_file_name_format(file_name: str) -> bool:
    """檢查檔案名稱是否包含 LivingRoom_DayX 或 Bedroom_DayX 格式"""
    pattern = r'(LivingRoom|Bedroom|Kitchen)_Day\d+'
    return bool(re.search(pattern, file_name))

# ============================
# 文件夾處理函式
# ----------------------------

def process_folder(folder_path: str, output_file: str, progress_callback: Optional[callable] = None) -> Optional[pd.DataFrame]:
    """處理整個文件夾的溫度數據並生成報告"""
    file_list = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and f != os.path.basename(output_file)]
    file_list.sort()
    
    # 檢查檔案名稱格式
    invalid_files = [f for f in file_list if not check_file_name_format(f)]
    if invalid_files:
        error_msg = "以下檔案名稱格式錯誤，應包含 Room_DayX 格式：\n" + "\n".join(invalid_files) + "\n\n是否跳過這些檔案並繼續處理？"
        if not messagebox.askyesno("檔案名稱錯誤", error_msg):
            return None
        file_list = [f for f in file_list if check_file_name_format(f)]
    
    results = []
    total_files = len(file_list)
    
    for idx, file_name in enumerate(file_list, 1):
        file_path = os.path.join(folder_path, file_name)
        room = extract_room_name(file_name)
        if room is None:
            continue
        
        result = process_file(file_path, room)
        if result[0] is not None:
            day_number = extract_day_number(file_name)
            if day_number is not None:
                results.append((room, day_number, result))
        
        if progress_callback:
            progress_callback(idx / total_files)
    
    results.sort(key=lambda x: (x[0], x[1]))
    
    final_result_data = {
        "Room_Day": [f"{room}_Day{day}" for room, day, _ in results],
        "Avg_Temperature": [result[0] for _, _, result in results],
        "Max_Temperature": [result[1] for _, _, result in results],
        "Time_Range": [result[2] for _, _, result in results],
        "Record_Count": [result[3] for _, _, result in results]
    }
    
    final_df = pd.DataFrame(final_result_data)
    _save_to_excel(final_df, output_file)
    
    return final_df

def _save_to_excel(df: pd.DataFrame, output_file: str) -> None:
    """保存數據到 Excel 文件，根據房間設置背景顏色，並自動調整欄寬"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Temperature Report')
        worksheet = writer.sheets['Temperature Report']
        
        # 根據距離值設置背景顏色
        for row in range(2, len(df) + 2):
            room_value = df.at[row - 2, 'Room_Day'].split('_')[0]
            fill = color_based_on_room(room_value)
            if fill:
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).fill = fill
        
        # 自動調整欄寬
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # 獲取欄位字母（如 'A'）
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # 加點餘裕
            worksheet.column_dimensions[column].width = adjusted_width


# ============================
# 主應用程式類別
# ----------------------------

class App:
    """智慧居家溫度彙整器的主應用程式類別"""
    
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self._setup_window()
        self._create_widgets()
        self.folder_path = ""
    
    def _setup_window(self) -> None:
        """設置主視窗的基本屬性 (暗色介面)"""
        self.root.title("智慧居家溫度彙整器")
        self.root.geometry("600x400")
        self.root.configure(bg='#2E2E2E')
        
        style = ttk.Style()
        style.theme_use('clam')
        dark_bg = '#2E2E2E'
        dark_fg = '#FFFFFF'
        for widget in ['TButton', 'TProgressbar', 'TFrame', 'TLabel']:
            style.configure(widget, background=dark_bg, foreground=dark_fg, relief=tk.FLAT, font=("Arial", 10))
        style.map("TButton", background=[('active', '#3E3E3E')])
    
    def _create_widgets(self) -> None:
        """創建和配置所有 UI 元件"""
        # 輸入框架
        input_frame = ttk.Frame(self.root, style='TFrame')
        input_frame.place(relx=0.5, rely=0.1, anchor=tk.CENTER)
        label = tk.Label(input_frame, text="選擇溫度數據文件夾:", font=("Arial", 12), bg='#2E2E2E', fg='#FFFFFF')
        label.pack(side=tk.LEFT, padx=5)
        self.select_button = tk.Button(input_frame, text="選擇文件夾", command=self.select_folder,
                                       bg="#4CAF50", fg="white", font=("Arial", 10), relief=tk.FLAT)
        self.select_button.pack(side=tk.LEFT, padx=5)
        
        # 路徑顯示標籤
        self.path_label = tk.Label(self.root, text="", font=("Arial", 10), bg='#2E2E2E', fg='#FFFFFF')
        self.path_label.place(relx=0.5, rely=0.2, anchor=tk.CENTER)
        
        # 進度條框架
        progress_frame = ttk.Frame(self.root, style='TFrame')
        progress_frame.place(relx=0.5, rely=0.4, anchor=tk.CENTER)
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, length=300, mode='determinate')
        self.progress.pack(pady=10)
        
        # 按鈕框架
        button_frame = ttk.Frame(self.root, style='TFrame')
        button_frame.place(relx=0.5, rely=0.6, anchor=tk.CENTER)
        self.start_button = tk.Button(button_frame, text="開始彙整", command=self.start_aggregation,
                                      state=tk.DISABLED, bg="#2196F3", fg="white", font=("Arial", 10), relief=tk.FLAT)
        self.start_button.pack(pady=5)
    
    def select_folder(self) -> None:
        """選擇要處理的文件夾"""
        self.folder_path = filedialog.askdirectory()
        if self.folder_path:
            self.start_button.config(state=tk.NORMAL)
            self.path_label.config(text=f"當前路徑: {self.folder_path}")
    
    def start_aggregation(self) -> None:
        """開始數據彙整流程"""
        if not self.folder_path:
            messagebox.showerror("錯誤", "請先選擇一個文件夾。")
            return
        
        output_file = os.path.join(self.folder_path, '溫度彙整報告.xlsx')
        self.progress['value'] = 0
        self.root.update_idletasks()
        
        try:
            result_df = process_folder(
                self.folder_path,
                output_file,
                lambda p: self._update_progress(p)
            )
            
            if result_df is not None:
                messagebox.showinfo("完成", "溫度彙整完成！")
            else:
                messagebox.showwarning("警告", "彙整未完成，可能檔案格式錯誤或無有效數據。")
        except Exception as e:
            messagebox.showerror("錯誤", f"發生錯誤: {str(e)}")
        
        self._reset_progress()
    
    def _update_progress(self, progress: float) -> None:
        """更新進度條"""
        self.progress['value'] = progress * 100
        self.root.update_idletasks()
    
    def _reset_progress(self) -> None:
        """重置進度條"""
        self.progress['value'] = 0
        self.root.update_idletasks()

# ============================
# 主程式入口
# ----------------------------

def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()